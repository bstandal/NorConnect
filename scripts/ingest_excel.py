#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

import psycopg
from dotenv import load_dotenv
from openpyxl import load_workbook
from psycopg.types.json import Jsonb

SHEET_TO_TABLE = {
    "Organisasjoner": "stg_excel_organisasjoner",
    "Datakilder": "stg_excel_datakilder",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Load Excel sheets into staging tables.")
    parser.add_argument(
        "--file",
        default=os.getenv("EXCEL_PATH"),
        help="Path to input .xlsx file (default: EXCEL_PATH env var).",
    )
    parser.add_argument(
        "--truncate",
        action="store_true",
        help="Truncate Excel staging tables before ingesting.",
    )
    return parser.parse_args()


def serialize_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, (int, float, bool, str)):
        return value
    return str(value)


def make_headers(raw_headers: list[Any]) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}

    for idx, raw in enumerate(raw_headers, start=1):
        base = str(raw).strip() if raw is not None else ""
        if not base:
            base = f"column_{idx}"

        count = seen.get(base, 0)
        seen[base] = count + 1
        if count:
            headers.append(f"{base}_{count + 1}")
        else:
            headers.append(base)

    return headers


def is_empty_row(values: tuple[Any, ...]) -> bool:
    for value in values:
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return False
    return True


def main() -> int:
    load_dotenv()
    args = parse_args()

    if not args.file:
        print("Input file required via --file or EXCEL_PATH.", file=sys.stderr)
        return 1

    file_path = Path(args.file).expanduser().resolve()
    if not file_path.exists():
        print(f"File not found: {file_path}", file=sys.stderr)
        return 1

    dsn = os.getenv("POSTGRES_DSN")
    if not dsn:
        print("POSTGRES_DSN is required.", file=sys.stderr)
        return 1

    workbook = load_workbook(file_path, data_only=False)

    with psycopg.connect(dsn) as conn:
        conn.autocommit = False

        run_id = conn.execute(
            """
            INSERT INTO ingest_run (source_name, input_path, status)
            VALUES (%s, %s, 'running')
            RETURNING id
            """,
            ("excel", str(file_path)),
        ).fetchone()[0]
        conn.commit()

        try:
            if args.truncate:
                for table in SHEET_TO_TABLE.values():
                    conn.execute(f"TRUNCATE TABLE {table}")
                conn.commit()

            total_rows = 0

            for sheet_name, table_name in SHEET_TO_TABLE.items():
                if sheet_name not in workbook.sheetnames:
                    print(f"warn: sheet not found, skipping: {sheet_name}")
                    continue

                ws = workbook[sheet_name]
                header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                if header_cells is None:
                    continue

                headers = make_headers(list(header_cells))

                inserted = 0
                for excel_row, row_values in enumerate(
                    ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True),
                    start=2,
                ):
                    if is_empty_row(row_values):
                        continue

                    payload = {
                        key: serialize_value(value)
                        for key, value in zip(headers, row_values)
                        if serialize_value(value) is not None
                    }
                    if not payload:
                        continue

                    conn.execute(
                        f"""
                        INSERT INTO {table_name}
                          (ingest_run_id, excel_source_path, excel_sheet, excel_row, row_payload)
                        VALUES (%s, %s, %s, %s, %s)
                        """,
                        (run_id, str(file_path), sheet_name, excel_row, Jsonb(payload)),
                    )
                    inserted += 1

                total_rows += inserted
                print(f"{sheet_name}: inserted {inserted} rows")

            conn.execute(
                """
                UPDATE ingest_run
                SET status = 'success', finished_at = now(), notes = %s
                WHERE id = %s
                """,
                (json.dumps({"rows_inserted": total_rows}), run_id),
            )
            conn.commit()
            print(f"Ingest complete. run_id={run_id}, total_rows={total_rows}")
            return 0

        except Exception as exc:  # noqa: BLE001
            conn.rollback()
            conn.execute(
                """
                UPDATE ingest_run
                SET status = 'failed', finished_at = now(), notes = %s
                WHERE id = %s
                """,
                (str(exc), run_id),
            )
            conn.commit()
            raise


if __name__ == "__main__":
    raise SystemExit(main())
